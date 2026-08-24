"""
Conditional formatting rule parsing (`conditionalFormatting`/`dxf` XML).

Verifies:
- `cellIs` rules (operator + formula(s)) resolve their dxfId against
  xl/styles.xml's <dxfs> into concrete font/fill overrides.
- Rule types with no dxfId (e.g. duplicateValues without a fill) still
  parse with format=None rather than being silently dropped.
- Multiple conditionalFormatting blocks / multiple rules are all captured.
- Files with no conditional formatting return an empty list, not an error.
"""

import os
import tempfile

import pytest
import streamxl
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle


@pytest.fixture
def workbook_with_conditional_formatting():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws["A1"] = 5
    ws["A2"] = 150
    ws["B1"] = "keep"
    ws["B2"] = "keep"

    red_bold = CellIsRule(
        operator="greaterThan",
        formula=["100"],
        font=Font(color="FFFF0000", bold=True),
        fill=PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
    )
    ws.conditional_formatting.add("A1:A10", red_bold)

    between_rule = CellIsRule(operator="between", formula=["1", "10"])
    ws.conditional_formatting.add("C1:C10", between_rule)

    # A rule type with no dxfId at all (duplicateValues with default styling)
    dup_rule = Rule(type="duplicateValues", priority=1)
    ws.conditional_formatting.add("B1:B10", dup_rule)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        yield tmp.name
    os.unlink(tmp.name)


def test_no_conditional_formatting_returns_empty_list(tmp_xlsx):
    assert streamxl.conditional_formats(tmp_xlsx) == []


def test_cell_is_rule_resolves_dxf(workbook_with_conditional_formatting):
    rules = streamxl.conditional_formats(workbook_with_conditional_formatting)
    by_sqref = {r["sqref"]: r for r in rules}

    rule = by_sqref["A1:A10"]
    assert rule["type"] == "cellIs"
    assert rule["operator"] == "greaterThan"
    assert rule["formulas"] == ["100"]
    assert rule["format"] is not None
    assert rule["format"]["font_color"] == "FFFF0000"
    assert rule["format"]["font_bold"] is True
    assert rule["format"]["fill_bg_color"] == "FFFFFF00"


def test_between_rule_has_two_formulas(workbook_with_conditional_formatting):
    rules = streamxl.conditional_formats(workbook_with_conditional_formatting)
    by_sqref = {r["sqref"]: r for r in rules}

    rule = by_sqref["C1:C10"]
    assert rule["operator"] == "between"
    assert rule["formulas"] == ["1", "10"]


def test_rule_without_dxf_id_has_none_format(workbook_with_conditional_formatting):
    rules = streamxl.conditional_formats(workbook_with_conditional_formatting)
    by_sqref = {r["sqref"]: r for r in rules}

    rule = by_sqref["B1:B10"]
    assert rule["type"] == "duplicateValues"
    assert rule["format"] is None


def test_all_rules_captured(workbook_with_conditional_formatting):
    rules = streamxl.conditional_formats(workbook_with_conditional_formatting)
    assert len(rules) == 3
    assert {r["sqref"] for r in rules} == {"A1:A10", "B1:B10", "C1:C10"}


def test_conditional_formats_respects_sheet_argument(workbook_with_conditional_formatting):
    rules_by_name = streamxl.conditional_formats(
        workbook_with_conditional_formatting, sheet="Data"
    )
    assert len(rules_by_name) == 3


def test_conditional_formats_raises_on_missing_file(tmp_path):
    missing = str(tmp_path / "does_not_exist.xlsx")
    with pytest.raises(Exception):
        streamxl.conditional_formats(missing)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
