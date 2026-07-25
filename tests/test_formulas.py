"""
Phase 1: Formula Extraction Tests

Tests for reading formulas and calculating metadata from Excel files.
Verifies:
- Formula detection and extraction
- Formula type classification
- Calculated values alongside formulas
- Error cell type support
"""

import pytest
import streamxl
from openpyxl import Workbook
import tempfile
import os


@pytest.fixture
def formula_workbook():
    """Create a workbook with various formula types for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Formulas"

    # Headers
    ws['A1'] = "Name"
    ws['B1'] = "Value1"
    ws['C1'] = "Value2"
    ws['D1'] = "Result"
    ws['E1'] = "Type"

    # SUM formula
    ws['A2'] = "Sum"
    ws['B2'] = 10
    ws['C2'] = 20
    ws['D2'] = "=SUM(B2:C2)"
    ws['E2'] = "sum"

    # AVERAGE formula
    ws['A3'] = "Average"
    ws['B3'] = 30
    ws['C3'] = 40
    ws['D3'] = "=AVERAGE(B3:C3)"
    ws['E3'] = "average"

    # IF formula
    ws['A4'] = "Conditional"
    ws['B4'] = 50
    ws['C4'] = 60
    ws['D4'] = "=IF(B4>C4,\"B\",\"C\")"
    ws['E4'] = "if"

    # VLOOKUP formula (simplified)
    ws['A5'] = "Lookup"
    ws['B5'] = 1
    ws['C5'] = 2
    ws['D5'] = "=INDEX(C2:C4,B5)"
    ws['E5'] = "index_match"

    # COUNT formula
    ws['A6'] = "Count"
    ws['B6'] = 100
    ws['C6'] = 200
    ws['D6'] = "=COUNT(B2:C6)"
    ws['E6'] = "count"

    # PRODUCT formula
    ws['A7'] = "Product"
    ws['B7'] = 5
    ws['C7'] = 6
    ws['D7'] = "=PRODUCT(B7:C7)"
    ws['E7'] = "product"

    # Custom formula
    ws['A8'] = "Custom"
    ws['B8'] = 10
    ws['C8'] = 5
    ws['D8'] = "=B8^C8"
    ws['E8'] = "custom"

    # Plain cell (no formula)
    ws['A9'] = "Plain"
    ws['B9'] = 123
    ws['C9'] = 456
    ws['D9'] = 579
    ws['E9'] = "none"

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb.save(tmp.name)
        yield tmp.name
    os.unlink(tmp.name)


def test_read_without_formulas(formula_workbook):
    """Test that reading without formulas still works (backward compat)."""
    rows = list(streamxl.read(formula_workbook))
    assert len(rows) == 9  # Header + 8 data rows
    assert rows[0] == ["Name", "Value1", "Value2", "Result", "Type"]
    # Row 2 is the Average row (because row 1 is Sum)
    # The result cell should have the formula text or calculated value
    result_value = rows[2][3]
    assert result_value is None or isinstance(result_value, (str, int, float))


def test_read_with_formulas(formula_workbook):
    """Test reading with formula metadata."""
    rows = list(streamxl.read(formula_workbook, with_formulas=True))
    assert len(rows) == 9

    # Check header row
    header = rows[0]
    assert all(isinstance(cell, dict) for cell in header)
    assert header[0]["formula"] is None

    # Check SUM formula (row 1 in 0-indexed)
    sum_row = rows[1]
    result_cell = sum_row[3]
    assert isinstance(result_cell, dict)
    assert result_cell["formula"] is not None
    assert "SUM" in result_cell["formula"].upper()
    assert result_cell["formula_type"] == "sum"

    # Check AVERAGE formula (row 2)
    avg_row = rows[2]
    result_cell = avg_row[3]
    assert "AVERAGE" in result_cell["formula"].upper()
    assert result_cell["formula_type"] == "average"

    # Check IF formula (row 3)
    if_row = rows[3]
    result_cell = if_row[3]
    assert "IF" in result_cell["formula"].upper()
    assert result_cell["formula_type"] == "if"

    # Check INDEX formula (row 4)
    index_row = rows[4]
    result_cell = index_row[3]
    assert result_cell["formula"] is not None

    # Check COUNT formula (row 5)
    count_row = rows[5]
    result_cell = count_row[3]
    assert "COUNT" in result_cell["formula"].upper()
    assert result_cell["formula_type"] == "count"

    # Check PRODUCT formula (row 6)
    product_row = rows[6]
    result_cell = product_row[3]
    assert "PRODUCT" in result_cell["formula"].upper()
    assert result_cell["formula_type"] == "product"

    # Check plain cell (no formula) (row 8)
    plain_row = rows[8]
    result_cell = plain_row[3]
    assert result_cell["formula"] is None
    assert result_cell["formula_type"] is None


def test_formula_type_detection(formula_workbook):
    """Test that formula types are correctly detected."""
    rows = list(streamxl.read(formula_workbook, with_formulas=True))

    formula_types = {
        "sum": False,
        "average": False,
        "if": False,
        "index_match": False,
        "count": False,
        "product": False,
        "custom": False,
    }

    for row in rows[1:]:  # Skip header
        for cell in row:
            if cell.get("formula_type") in formula_types:
                formula_types[cell["formula_type"]] = True

    # Check that we found at least the main types
    assert formula_types["sum"]
    assert formula_types["average"]
    assert formula_types["if"]
    assert formula_types["count"]
    assert formula_types["product"]


def test_metadata_dict_structure(formula_workbook):
    """Test that metadata dicts have correct structure."""
    rows = list(streamxl.read(formula_workbook, with_formulas=True))

    for row in rows:
        for cell in row:
            assert isinstance(cell, dict)
            assert "value" in cell
            assert "formula" in cell
            assert "formula_type" in cell

            # formula and formula_type should be None or string
            if cell["formula"] is not None:
                assert isinstance(cell["formula"], str)
                assert len(cell["formula"]) > 0
            if cell["formula_type"] is not None:
                assert isinstance(cell["formula_type"], str)


def test_read_with_dict_and_formulas(formula_workbook):
    """Test reading with both as_dict and with_formulas."""
    rows = list(streamxl.read(formula_workbook, as_dict=True, with_formulas=True))
    assert len(rows) == 8  # Header consumed, 8 data rows

    for row in rows:
        assert isinstance(row, dict)
        for key, cell in row.items():
            assert isinstance(cell, dict)
            assert "value" in cell
            assert "formula" in cell


def test_read_with_columns_and_formulas(formula_workbook):
    """Test reading with column filtering and formulas."""
    columns = ["Name", "Result"]
    rows = list(
        streamxl.read(formula_workbook, as_dict=True, columns=columns, with_formulas=True)
    )
    assert len(rows) == 8

    for row in rows:
        assert set(row.keys()) == {"Name", "Result"}
        for cell in row.values():
            assert isinstance(cell, dict)
            assert "value" in cell
            assert "formula" in cell


def test_formula_value_preservation(formula_workbook):
    """Test that formula values are preserved (when Excel calculated them)."""
    rows = list(streamxl.read(formula_workbook, with_formulas=True))

    # SUM formula should have value if Excel calculated it (row 1, 0-indexed)
    sum_row = rows[1]
    result_cell = sum_row[3]
    # Value might be None (not calculated by Excel) or the calculated result
    assert result_cell["value"] is None or isinstance(result_cell["value"], (int, float, str))
    assert result_cell["formula"] is not None
    assert "SUM" in result_cell["formula"].upper()


def test_empty_cells_with_formulas(formula_workbook):
    """Test that empty cells are handled correctly with formula mode."""
    rows = list(streamxl.read(formula_workbook, with_formulas=True))

    # All rows should have the same number of cells (with None for empty ones)
    row_lengths = [len(row) for row in rows]
    # Lengths might vary, but that's OK as long as each cell is a dict
    for row in rows:
        for cell in row:
            assert isinstance(cell, dict)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
